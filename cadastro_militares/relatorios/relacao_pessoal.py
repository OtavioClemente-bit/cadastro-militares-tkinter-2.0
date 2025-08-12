# relatorios/relacao_pessoal.py
# Gera a "Relação Pessoal" em Excel
# Requisitos: openpyxl (pip install openpyxl)

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# DB
from database.db import buscar_todos

# ------------------------------
#  Configurações visuais
# ------------------------------
HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")  # cinza claro
HEADER_FONT = Font(bold=True, color="000000")         # preto
ORDER_FONT  = Font(color="666666")                    # números em cinza
BODY_FONT   = Font(color="000000")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
THIN   = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ------------------------------
#  Hierarquia e abreviações
# ------------------------------
# Ordem pedida (do mais antigo para o mais moderno), conforme usuário
ORDEM_HIERARQUIA = [
    "Capitão",
    "1º Tenente",
    "2º Tenente",
    "Subtenente",
    "3º Sargento",
    "2º Sargento",
    "1º Sargento",
    "Aspirante",
    "Soldado Efetivo Profissional",
    "Soldado Efetivo Variável",
    "Cabo Efetivo Profissional",
]
ORD_IDX = {p: i for i, p in enumerate(ORDEM_HIERARQUIA)}

ABREV = {
    "Capitão": "CAP",
    "1º Tenente": "1º TEN",
    "2º Tenente": "2º TEN",
    "Subtenente": "ST",
    "1º Sargento": "1º SGT",
    "2º Sargento": "2º SGT",
    "3º Sargento": "3º SGT",
    "Aspirante": "ASP",
    "Cabo Efetivo Profissional": "CB EF PROFL",
    "Soldado Efetivo Profissional": "SD EF PROFL",
    "Soldado Efetivo Variável": "SD EF VRV",
}

# ------------------------------
#  Helpers
# ------------------------------
def _digits_only(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())

def _abbr_posto(posto: str) -> str:
    return ABREV.get(str(posto), str(posto).upper())

def _ord_key(reg):
    """
    reg é a tupla/linha retornada por buscar_todos():
    (0)ID, (1)Posto, (2)Nome, (3)NomeGuerra, (4)CPF, (5)PREC, ...
    """
    posto = reg[1]
    nome  = str(reg[2] or "")
    return (ORD_IDX.get(posto, 999), nome.lower())

def _default_path() -> str:
    desk = os.path.join(os.path.expanduser("~"), "Desktop")
    base = "Relacao_Pessoal.xlsx"
    return os.path.join(desk if os.path.isdir(desk) else os.getcwd(), base)

# ------------------------------
#  Função principal
# ------------------------------
def gerar_relacao_pessoal(parent=None, caminho_arquivo: str | None = None):
    """
    Gera a planilha Excel com as colunas:
      NR ORDEM | PREC-CP | P/G | NOME COMPLETO | CPF
    - Cabeçalho cinza
    - NR ORDEM em fonte cinza
    - Nomes e P/G em MAIÚSCULO
    - CPF só com dígitos (sem pontos/traço)
    - Ordenado por hierarquia especificada e depois por nome
    """
    # Carrega/ordena dados
    regs = sorted(buscar_todos(), key=_ord_key)

    # Monta workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relação"

    # Cabeçalho
    headers = ["NR ORDEM", "PREC-CP", "P/G", "NOME COMPLETO", "CPF"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER_ALL

    # Linhas
    ordem = 1
    row_idx = 2
    for it in regs:
        posto = it[1]
        nome  = (it[2] or "").upper().strip()
        prec  = _digits_only(it[5])
        cpf   = _digits_only(it[4])
        pg    = _abbr_posto(posto)

        ws.cell(row=row_idx, column=1, value=ordem)        # NR ORDEM
        ws.cell(row=row_idx, column=2, value=prec)         # PREC-CP
        ws.cell(row=row_idx, column=3, value=pg)           # P/G
        ws.cell(row=row_idx, column=4, value=nome)         # NOME COMPLETO
        ws.cell(row=row_idx, column=5, value=cpf)          # CPF

        # Estilos por coluna
        c1 = ws.cell(row=row_idx, column=1); c1.alignment = CENTER; c1.font = ORDER_FONT
        c2 = ws.cell(row=row_idx, column=2); c2.alignment = CENTER; c2.font = BODY_FONT
        c3 = ws.cell(row=row_idx, column=3); c3.alignment = CENTER; c3.font = BODY_FONT
        c4 = ws.cell(row=row_idx, column=4); c4.alignment = LEFT;   c4.font = BODY_FONT
        c5 = ws.cell(row=row_idx, column=5); c5.alignment = CENTER; c5.font = BODY_FONT

        # Borda em todas as células
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = BORDER_ALL

        # Garante formato TEXTO para PREC e CPF (preserva zeros à esquerda)
        ws.cell(row=row_idx, column=2).number_format = "@"
        ws.cell(row=row_idx, column=5).number_format = "@"

        ordem += 1
        row_idx += 1

    # Largura das colunas
    widths = {
        1: 12,   # NR ORDEM
        2: 16,   # PREC-CP
        3: 16,   # P/G
        4: 54,   # NOME COMPLETO
        5: 16,   # CPF
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Congela cabeçalho
    ws.freeze_panes = "A2"

    # Rodapé opcional (data/hora geração) como comentário simples na célula A1
    ws.cell(row=1, column=1).comment = None  # garante limpo
    ws.cell(row=1, column=1).comment = None  # (mantém simples)

    # Salvar
    if caminho_arquivo is None:
        # Se tiver um parent Tk, pergunta onde salvar; senão, salva no Desktop
        try:
            if parent is not None:
                from tkinter import filedialog
                caminho_arquivo = filedialog.asksaveasfilename(
                    parent=parent,
                    defaultextension=".xlsx",
                    filetypes=[("Planilha Excel", "*.xlsx")],
                    initialfile="Relacao_Pessoal.xlsx",
                    title="Salvar Relação Pessoal"
                )
                if not caminho_arquivo:
                    return  # cancelou
            else:
                caminho_arquivo = _default_path()
        except Exception:
            caminho_arquivo = _default_path()

    wb.save(caminho_arquivo)

    # Feedback
    try:
        if parent is not None:
            from tkinter import messagebox
            messagebox.showinfo("OK", f"Relação gerada com sucesso:\n{caminho_arquivo}", parent=parent)
    except Exception:
        pass

    return caminho_arquivo
