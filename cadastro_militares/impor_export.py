# impor_export.py — importador com mapeamento único/estrito e correções de NG, datas, ano e AT
import re
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.utils.datetime import from_excel as excel_date_from_serial, WINDOWS_EPOCH
from tkinter import filedialog, messagebox

from database.db import (
    buscar_todos, inserir_militar, atualizar_militar,
    listar_postos, inserir_posto
)

colunas = [
    "ID", "Posto", "Nome", "Nome de Guerra", "CPF", "PREC-CP", "IDT", "Banco",
    "Agência", "Conta", "Foto", "Ano de Formação", "Data de Nascimento", "Data de Praça",
    "Endereço", "CEP", "Recebe Pré Escolar", "Valor Pré Escolar",
    "Recebe Auxílio Transporte", "Valor Auxílio Transporte", "Possui PNR"
]

PG_ABREV_TO_FULL = {
    "CAP": "Capitão",
    "1º TEN": "1º Tenente", "1° TEN": "1º Tenente",
    "2º TEN": "2º Tenente", "2° TEN": "2º Tenente",
    "ST": "Subtenente",
    "1º SGT": "1º Sargento", "1° SGT": "1º Sargento",
    "2º SGT": "2º Sargento", "2° SGT": "2º Sargento",
    "3º SGT": "3º Sargento", "3° SGT": "3º Sargento",
    "ASP": "Aspirante",
    "SD EF PROFL": "Soldado Efetivo Profissional",
    "SD EF VRV": "Soldado Efetivo Variável",
    "CB EF PROFL": "Cabo Efetivo Profissional",
}

# ---------------- helpers ----------------
def _so_digitos(s):
    return re.sub(r"\D+", "", str(s or ""))

def _strip(s):
    return str(s or "").strip()

def _norm_header(s: str) -> str:
    t = "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn")
    t = re.sub(r"[^A-Za-z0-9]+", " ", t).strip().upper()
    return re.sub(r"\s+", " ", t)

def _primeiro_nome(nome_original: str) -> str:
    s = _strip(nome_original)
    if not s:
        return ""
    # pega o primeiro token só de letras
    m = re.match(r"[A-Za-zÀ-ÖØ-öø-ÿ]+", s)
    return (m.group(0).capitalize() if m else s.split()[0].capitalize())

def _parse_money(x):
    """Normaliza valores em '1234.56' (string). Aceita número, '2.866,66', 'R$ 286,66' etc."""
    if x is None or str(x).strip() == "":
        return ""
    if isinstance(x, (int, float, Decimal)):
        try:
            v = Decimal(str(x))
            return "0" if v <= 0 else f"{v:.2f}"
        except InvalidOperation:
            return ""
    s = str(x).strip()
    s = re.sub(r"[^\d,.\-]", "", s)   # remove R$, espaços, etc
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        v = Decimal(s)
        return "0" if v <= 0 else f"{v:.2f}"
    except InvalidOperation:
        return ""

PT_MONTHS = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12
}

def _parse_date(x, epoch=WINDOWS_EPOCH):
    """
    Aceita: datetime, serial Excel (número ou string), 'dd/mm/aaaa', 'dd/mm/aa',
            'yyyy-mm-dd', 'dd-mm-yy', '05 JUN 25' (meses PT).
    Retorna 'dd/mm/aaaa' ou ''.
    """
    if x is None or (isinstance(x, str) and x.strip() == ""):
        return ""
    if isinstance(x, datetime):
        return x.strftime("%d/%m/%Y")
    if isinstance(x, (int, float)) or (isinstance(x, str) and x.isdigit()):
        try:
            dt = excel_date_from_serial(float(x), epoch)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
    s = str(x).strip()
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})\s+(\d{2,4})$", s, flags=re.I)
    if m:
        d = int(m.group(1))
        mm_txt = m.group(2).upper()
        a = int(m.group(3))
        if a < 100: a += 2000
        mm = PT_MONTHS.get(mm_txt)
        if mm:
            try:
                return datetime(a, mm, d).strftime("%d/%m/%Y")
            except Exception:
                pass
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except Exception:
            continue
    return ""

def _ano_from_cell(x):
    """Extrai SÓ o ano: 2018, '2018', 2018.0, 'Form. 2018' -> '2018'."""
    if x is None or str(x).strip() == "":
        return ""
    if isinstance(x, (int, float)):
        try:
            return str(int(x))
        except Exception:
            pass
    s = str(x)
    m = re.search(r"(19|20)\d{2}", s)
    return m.group(0) if m else ""

def _sim_nao_from_cell(x):
    s = _strip(x).upper()
    if s in {"X", "NAO", "NÃO", "N", "NAO TEM", "NADA", "N/T", "NA"}:
        return "Não"
    if s in {"SIM", "S"}:
        return "Sim"
    return ""  

def _garante_posto(posto_full):
    if not posto_full:
        return
    try:
        existentes = set(listar_postos() or [])
        if posto_full not in existentes:
            inserir_posto(posto_full)
    except Exception:
        pass

# ---------------- export ----------------
def exportar_para_excel(janela):
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Salvar como"
    )
    if not arquivo:
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Militares"
        for i, col in enumerate(colunas, 1):
            ws.cell(row=1, column=i).value = col
        for i, dado in enumerate(buscar_todos(), 2):
            for j, valor in enumerate(dado, 1):
                ws.cell(row=i, column=j).value = valor
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{arquivo}", parent=janela)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar: {e}", parent=janela)

# ---------------- import (único/estrito) ----------------
def importar_de_excel(janela, carregar_militares_callback):
    caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not caminho:
        return
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb.active
        EPOCH = getattr(wb, "epoch", WINDOWS_EPOCH)

        # localizar cabeçalho (primeira linha não vazia nas 10 primeiras)
        header = None
        header_row_idx = 1
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any(cell not in (None, "") for cell in row):
                header = [str(c or "").strip() for c in row]
                header_row_idx = i
                break
        if not header:
            messagebox.showerror("Erro", "Não encontrei cabeçalho na planilha.", parent=janela)
            return

        headers_norm = [_norm_header(h) for h in header]
        used_cols = set()

        def find_col(aliases, required_tokens=None):
            """
            Procura coluna respeitando:
            - não pode já estar usada;
            - match exato > contém;
            - se required_tokens for dado, exige ao menos um token presente.
            """
            def ok_tokens(hn):
                if not required_tokens: return True
                tokens = set(hn.split())
                return any(tok in tokens for tok in required_tokens)

            # exato
            for i, hn in enumerate(headers_norm):
                if i in used_cols: continue
                if any(hn == _norm_header(al) for al in aliases) and ok_tokens(hn):
                    used_cols.add(i); return i
            # contém
            for i, hn in enumerate(headers_norm):
                if i in used_cols: continue
                if any(_norm_header(al) in hn or hn in _norm_header(al) for al in aliases) and ok_tokens(hn):
                    used_cols.add(i); return i
            return None

        colmap = {
            "pg":   find_col(["P G","PG","POSTO GRADUACAO","POSTO E GRADUACAO","POSTO","GRADUACAO","P/G"], {"PG","POSTO","GRADUACAO"}),
            "nome": find_col(["NOME","NOME COMPLETO"], {"NOME"}),
            # nome de guerra EXIGE token GUERRA/NG
            "nome_guerra": find_col(["NOME DE GUERRA","GUERRA","NG"], {"GUERRA","NG"}),
            "cpf":  find_col(["CPF"], {"CPF"}),
            "prec": find_col(["PREC","PREC CP","PREC-CP","PREC CP.","PRECCP"], {"PREC"}),
            "idt":  find_col(["IDT","IDENTIDADE","IDT MILITAR","IDENT MILITAR","IDENTIDADE MILITAR"], {"IDT","IDENTIDADE"}),
            "banco": find_col(["BANCO"], {"BANCO"}),
            "agencia": find_col(["AGENCIA","AGÊNCIA","AG."], {"AGENCIA","AG"}),
            "conta": find_col(["CONTA","CONTA CORRENTE","C/C"], {"CONTA"}),
            # ano exige token ANO/FORM
            "ano": find_col(["ANO","ANO DE FORMACAO","ANO FORMACAO","ANO FORM","ANO DA FORMACAO","ANO FORM.","FORMACAO","ANO FORMATURA"], {"ANO","FORMACAO","FORM"}),
            # nascimento exige token NASC/DN
            "nasc": find_col(["DATA DE NASCIMENTO","NASCIMENTO","DT NASC","DN","DT NASCIMENTO","ANIVERSARIO","ANIVERSÁRIO","DT. NASC."], {"NASC","DN"}),
            # praça exige token PRACA/PRAÇA
            "praca": find_col(["DATA DE PRACA","DT PRACA","PRACA","DATA PRACA","DATA DE PRAÇA","DT PRAÇA"], {"PRACA","PRACA","PRAÇA"}),
            "endereco": find_col(["ENDERECO","ENDEREÇO"], {"ENDERECO"}),
            "cep": find_col(["CEP"], {"CEP"}),
            # pré-escolar exige token PRE
            "valor_pre": find_col(["VALOR PRE ESCOLAR","PRE ESCOLAR","VALOR PRE","PRE"], {"PRE"}),
            "rec_pre": find_col(["RECEBE PRE ESCOLAR","REC PRE","PRE ESCOLAR S N","PRE ESCOLAR SN","PRE ESCOLAR (S/N)"], {"PRE"}),
            # AT exige token AT/TRANSP
            "valor_at": find_col(["VALOR AUXILIO TRANSPORTE","VALOR AT","AUX TRANSPORTE VALOR","AUX TRANSPORTE (R$)","AT (R$)","VALOR A T","VALOR A/T","VALOR AUX. TRANSPORTE"], {"AT","TRANSPORTE","AUX"}),
            "rec_at": find_col(["AUXILIO TRANSPORTE","AUX TRANSPORTE","RECEBE AT","AT S N","AT SN","A/T","AUX. TRANSPORTE"], {"AT","TRANSPORTE","AUX"}),
            "pnr": find_col(["PNR","POSSUI PNR","IMOVEL FUNCIONAL","IMÓVEL FUNCIONAL"], {"PNR","IMOVEL","IMÓVEL"}),
            "foto": find_col(["FOTO","CAMINHO FOTO","FOTO ARQUIVO"], {"FOTO"}),
        }

        existentes = buscar_todos()
        by_prec = {str(reg[5] or ""): reg for reg in existentes}   # 5 = PREC-CP
        by_cpf  = {str(reg[4] or ""): reg for reg in existentes}   # 4 = CPF

        inseridos = 0
        atualizados = 0
        ignorados = 0

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or all((c is None or str(c).strip() == "") for c in row):
                continue

            def get(colkey):
                ci = colmap.get(colkey)
                return row[ci] if ci is not None and ci < len(row) else None

            pg_raw   = get("pg")
            nome_raw = get("nome")
            ng_raw   = get("nome_guerra")
            cpf_raw  = get("cpf")
            prec_raw = get("prec")
            idt_raw  = get("idt")
            banco    = _strip(get("banco"))
            agencia  = _strip(get("agencia"))
            conta    = _strip(get("conta"))
            foto     = _strip(get("foto"))
            ano_raw  = get("ano")
            nasc_raw = get("nasc")
            praca_raw= get("praca")
            endereco = _strip(get("endereco"))
            cep_raw  = get("cep")
            vpre_raw = get("valor_pre")
            rpre_raw = get("rec_pre")
            vat_raw  = get("valor_at")
            rat_raw  = get("rec_at")
            pnr_raw  = get("pnr")

            # normalizações
            nome_upper = _strip(nome_raw).upper()
            if not nome_upper:
                ignorados += 1
                continue

            # NG: usa coluna específica; se vazia, primeiro nome do original (não do upper)
            nome_guerra = _strip(ng_raw) or _primeiro_nome(_strip(nome_raw))

            cpf  = _so_digitos(cpf_raw)
            prec = _so_digitos(prec_raw)
            idt  = _so_digitos(idt_raw)
            cep  = _so_digitos(cep_raw)

            pg_abrev = _strip(pg_raw).upper()
            posto_full = PG_ABREV_TO_FULL.get(pg_abrev, pg_abrev.title() if pg_abrev else "")
            _garante_posto(posto_full)

            ano = _ano_from_cell(ano_raw)
            data_nasc  = _parse_date(nasc_raw, EPOCH)
            data_praca = _parse_date(praca_raw, EPOCH)

            # Pré-escolar — só "Sim" se houver valor > 0 OU coluna explícita indicar Sim.
            valor_pre = _parse_money(vpre_raw)
            rec_pre   = _sim_nao_from_cell(rpre_raw)
            if rec_pre == "Não":
                valor_pre = "0"
            elif rec_pre == "Sim":
                if valor_pre in ("", "0"):
                    valor_pre = "0"
            else:
                # sem coluna explícita: deduz do valor (se vazio, Não)
                rec_pre = "Sim" if valor_pre not in ("", "0") else "Não"
                if rec_pre == "Não": valor_pre = "0"

            # Auxílio-Transporte — sua regra
            valor_at = _parse_money(vat_raw)
            if _strip(rat_raw).upper() == "X":
                rec_at = "Não"; valor_at = "0"
            else:
                if valor_at not in ("", "0"):
                    rec_at = "Sim"
                else:
                    rec_at = "Não"; valor_at = "0"

            pnr = _sim_nao_from_cell(pnr_raw) or ""

            # upsert por PREC (fallback CPF)
            atual = None; rid = None
            if prec and prec in by_prec:
                atual = by_prec[prec]; rid = atual[0]
            elif cpf and cpf in by_cpf:
                atual = by_cpf[cpf]; rid = atual[0]

            def pick(novo, antigo):
                nv = _strip(novo)
                return nv if nv not in ("", None) else _strip(antigo)

            if atual:
                (_, a_posto, a_nome, a_ng, a_cpf, a_prec, a_idt,
                 a_banco, a_ag, a_cc, a_foto, a_ano, a_nasc, a_praca,
                 a_end, a_cep, a_recpre, a_vpre, a_recat, a_vat, a_pnr) = atual

                novo_reg = (
                    pick(posto_full, a_posto),
                    pick(nome_upper, a_nome),
                    pick(nome_guerra, a_ng),
                    pick(cpf, a_cpf),
                    pick(prec, a_prec),
                    pick(idt, a_idt),
                    pick(banco, a_banco),
                    pick(agencia, a_ag),
                    pick(conta, a_cc),
                    pick(foto, a_foto),
                    pick(ano, a_ano),
                    pick(data_nasc, a_nasc),
                    pick(data_praca, a_praca),
                    pick(endereco, a_end),
                    pick(cep, a_cep),
                    (rec_pre or a_recpre or "Não"),
                    (valor_pre if valor_pre not in ("", None) else (a_vpre or "0")),
                    rec_at,
                    valor_at,
                    (pnr or a_pnr or "Não"),
                )
                atualizar_militar(rid, novo_reg)
                atualizados += 1
            else:
                novo_reg = (
                    posto_full, nome_upper, nome_guerra, cpf, prec, idt,
                    banco, agencia, conta, foto,
                    ano, data_nasc, data_praca, endereco, cep,
                    rec_pre, valor_pre,
                    rec_at, valor_at,
                    pnr or "Não",
                )
                inserir_militar(novo_reg)
                inseridos += 1

        carregar_militares_callback()

        # resumo de mapeamento
        wanted = {
            "pg":"Posto/Graduação","nome":"Nome","nome_guerra":"Nome de Guerra","cpf":"CPF","prec":"PREC-CP","idt":"IDT",
            "banco":"Banco","agencia":"Agência","conta":"Conta","ano":"Ano de Formação",
            "nasc":"Data de Nascimento","praca":"Data de Praça","endereco":"Endereço","cep":"CEP",
            "valor_pre":"Valor Pré-Escolar","rec_pre":"Recebe Pré-Escolar",
            "valor_at":"Valor Auxílio Transporte","rec_at":"Recebe Auxílio Transporte",
            "pnr":"PNR","foto":"Foto"
        }
        resumo_map = []
        for k, desc in wanted.items():
            ci = colmap.get(k)
            if ci is None:
                resumo_map.append(f"– {desc}: NÃO ENCONTRADO")
            else:
                resumo_map.append(f"– {desc}: coluna {ci+1} ({header[ci]})")

        messagebox.showinfo(
            "Importação concluída",
            "Inseridos: {}\nAtualizados: {}\nIgnorados: {}\n\nMapeamento de colunas:\n{}".format(
                inseridos, atualizados, ignorados, "\n".join(resumo_map)
            ),
            parent=janela
        )

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao importar: {e}", parent=janela)
